import RPi.GPIO as GPIO
import time 
from datetime import datetime, date, timedelta, time
import openpyxl
import threading
import random

################################################################# DEFINICION DE PINES #################################################################

#LED ROJO COCHE
gpio.setmode(gpio.BOARD)
gpio.setup(13, gpio.OUT)
gpio.output(13, False)

#LED AMBAR COCHE
gpio.setmode(gpio.BOARD)
gpio.setup(19, gpio.OUT)
gpio.output(19, False)

#LED VERDE COCHE
gpio.setmode(gpio.BOARD)
gpio.setup(26, gpio.OUT)
gpio.output(26, False)

#LED ROJO PEATON
gpio.setmode(gpio.BOARD)
gpio.setup(5, gpio.OUT)
gpio.output(5, False)

#LED VERDE PEATON
gpio.setmode(gpio.BOARD)
gpio.setup(6, gpio.OUT)
gpio.output(6, False)

#PULSADOR SEMAFORO
gpio.setmode(gpio.BOARD)
gpio.setup(25, gpio.IN)

#INTERRUPTOR CONTROLADOR
gpio.setmode(gpio.BOARD)
gpio.setup(12, gpio.IN)

################################################################# DEFINICION DE CLASES Y FUNCIONES #################################################################


# FUNCIONES SEMAFORO
def cochesRojo(tiempo_peatones):
    gpio.output(13, True) # rojo coches
    gpio.output(19, False) # ambar coches
    gpio.output(26, False) # verde coches
    gpio.output(5, False) # rojo peaton
    gpio.output(6, True) # verde peaton

    time.sleep(tiempo_peatones)

def cochesAmbar():
    gpio.output(13, False) # rojo coches
    gpio.output(19, True) # ambar coches
    gpio.output(26, False) # verde coches
    gpio.output(5, True) # rojo peaton
    gpio.output(6, False) # verde peaton
    time.sleep(5)

def cochesVerde(tiempo_coches):
    gpio.output(13, False) # rojo coches
    gpio.output(19, False) # ambar coches
    gpio.output(26, False) # verde coches
    gpio.output(5, False) # rojo peaton
    gpio.output(6, False) # verde peaton

    time.sleep(tiempo_coches)

def transicionTodoRojo():
    gpio.output(13, True) # rojo coches
    gpio.output(19, False) # ambar coches
    gpio.output(26, False) # verde coches
    gpio.output(5, True) # rojo peaton
    gpio.output(6, False) # verde peaton

    time.sleep(3)

def noche():
    gpio.output(19, True) # ambar coches
    time.sleep(1)
    gpio.output(19, False) # ambar coches
    time.sleep(1)


# FUNCIÓN SUSTITUYENTE DE SENSORES DE AFLUENCIA
def sensores_afluencia():
    tiempos = []
    segundos_repartir = 80
    ##### HORA ACTUAL #####
    hora_actual = datetime.now()
    hora_actual = time(hora_actual.hour, hora_actual.minute, hora_actual.second)
    print(hora_actual)

    ##### HORARIOS #####
    hora_6 = time(6, 0, 0)
    hora_10 = time(10, 0, 0)
    hora_13 = time(13, 0, 0)
    hora_15 = time(15, 0, 0)
    hora_17 = time(17, 0, 0)
    hora_21 = time(21, 0, 0)
    hora_0 = time(23, 59, 59)

    afluencia_coches = 1
    afluencia_gente = 1

    # Para simular una situación real, dependiendo de la hora cambian las afluencias de coches y personas
    if hora_actual > hora_6 and hora_actual <= hora_10:
            afluencia_coches = random.randint(768, 1023)
            afluencia_gente = random.randint(0, 255)

    if hora_actual > hora_10 and hora_actual <= hora_13:
            afluencia_coches = random.randint(255, 768)
            afluencia_gente = random.randint(768, 1023)

    if hora_actual > hora_13 and hora_actual <= hora_15:
            afluencia_coches = random.randint(768, 1023)
            afluencia_gente = random.randint(255, 768)

    if hora_actual > hora_15 and hora_actual <= hora_17:
            afluencia_coches = random.randint(0, 255)
            afluencia_gente = random.randint(0, 255)

    if hora_actual > hora_17 and hora_actual <= hora_21:
            afluencia_coches = random.randint(768, 1023)
            afluencia_gente = random.randint(768, 1023)

    if hora_actual > hora_21 and hora_actual <= hora_0:
            afluencia_coches = random.randint(768, 1023)
            afluencia_gente = random.randint(255, 768)

    print(afluencia_coches)
    print(afluencia_gente)

    suma_afluencias = afluencia_coches + afluencia_gente

    porcentage_afluencia_coche = afluencia_coches / suma_afluencias

    tiempo_coches = segundos_repartir * porcentage_afluencia_coche
    tiempo_gente = segundos_repartir - tiempo_coches
    tiempos.append(tiempo_coches)
    tiempos.append(tiempo_gente)
    tiempos.append(afluencia_coches)
    tiempos.append(afluencia_gente)

    return tiempos


# FUNCIONES BASE DE DATOSEXCEL
def introducir_BBDD(afluencia_peatones, afluencia_coches, tiempo_rojo_coches, tiempo_ambar_coches, tiempo_verde_coches,tiempo_rojo_peatones, tiempo_verde_peatones):
    wb = openpyxl.load_workbook(r'C:\Users\Industria 4.0\Desktop\ProyectoFinal\ExcelDB.xlsx')
    sheet = wb.active

    # Encontrar las últimas columnas y filas con datos
    rows = sheet.max_row # Solo necesitamos saber la fila

    # Añadir datos
    sheet.cell(row=rows+1, column=1).value = datetime.now().strftime('%Y/%m/%d %H:%M:%S')

    sheet.cell(row=rows+1, column=2).value = tiempo_rojo_coches
    sheet.cell(row=rows+1, column=3).value = tiempo_ambar_coches
    sheet.cell(row=rows+1, column=4).value = tiempo_verde_coches
    sheet.cell(row=rows+1, column=5).value = tiempo_rojo_peatones
    sheet.cell(row=rows+1, column=6).value = tiempo_verde_peatones

    sheet.cell(row=rows+1, column=7).value = afluencia_coches
    sheet.cell(row=rows+1, column=8).value = afluencia_gente

    wb.save(r"C:\Users\Industria 4.0\Desktop\ProyectoFinal\ExcelDB.xlsx")


################################################################# FUNCIONAMIENTO DEL PROGRAMA #################################################################


# Semaforo
estado = True
limite_noche = time(0, 0, 0)
limite_madrugada = time(6, 0, 0)

while(estado == True):
    hora_actual = datetime.now()
    hora_actual = time(hora_actual.hour, hora_actual.minute, hora_actual.second)

    # Si son entre las 12 y las 5 se pone en modo noche
    if hora_actual > hora_limite_noche and hora_actual < hora_limite_madrugada:
        noche()

    tiempos = sensores_afluencia() # llamamos a la función que hace de sensor de afluencia, nos devolverá un array con los tiempos del coche y gente
    tiempo_coches = tiempos(0) # extraer del array retornado el primer valor
    tiempo_peatones = tiempos(1) # extraer del array retornado el segundo valor
    afluencia_coches = tiempos(2) # extraer del array retornado el segundo valor
    afluencia_peatones = tiempos(3) # extraer del array retornado el segundo valor
    
    # Completar tiempos
    tiempo_rojo_coches = tiempo_peatones + 10 # minimo 10 segundos
    tiempo_ambar_coches = 5
    tiempo_verde_coches = tiempo_coches
    tiempo_rojo_peatones = tiempo_coches + 6 # minimo 6 segundos
    tiempo_verde_peatones = tiempo_peatones

    # Llamar a las funciones para cambiar los colores del semaforo
    cochesRojo(tiempo_peatones)
    cochesAmbar()
    transicionTodoRojo()
    cochesVerde(tiempo_coches)
    transicionTodoRojo()
    
    #Introducir datos en el Excel
    introducir_BBDD(afluencia_peatones, afluencia_coches, tiempo_rojo_coches, tiempo_ambar_coches, tiempo_verde_coches,tiempo_rojo_peatones, tiempo_verde_peatones)